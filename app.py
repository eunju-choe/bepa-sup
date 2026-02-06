from flask import Flask, render_template, request, send_file # type: ignore
import pandas as pd # type: ignore
import os
import re
import numpy as np
import warnings
import openpyxl
warnings.filterwarnings('ignore')

app = Flask(__name__)

# 업로드 및 처리 폴더 설정
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

@app.route('/')
def index():
    return render_template('index.html')

"""
=============== 관내여비 담당자 서포터 기능 ===============
"""
@app.route('/trip')
def trip_index():
    return render_template('trip_index.html')

@app.route('/trip/upload', methods=['POST'])
def upload_and_process_trip_files():
    if 'trip_file' not in request.files or 'tag_file' not in request.files:
        return 'No file part'
    
    trip_file = request.files['trip_file']
    tag_file = request.files['tag_file']
    
    if trip_file.filename == '' or tag_file.filename == '':
        return 'No selected file'
    
    # 경로 설정 및 저장
    trip_path = os.path.join(app.config['UPLOAD_FOLDER'], 'trip_all.xlsx')
    tag_path = os.path.join(app.config['UPLOAD_FOLDER'], 'tag_all.xlsx')
    trip_file.save(trip_path)
    tag_file.save(tag_path)
    
    try:
        # 1. 데이터 로드 및 전처리
        df_trip = pd.read_excel(trip_path, header=1)
        col_tmp = pd.read_excel(trip_path, nrows=0).columns
        col_tmp = col_tmp[:col_tmp.get_loc('출장기간') - 1]
        df_trip.columns.values[:len(col_tmp)+1] = pd.read_excel(trip_path, nrows=0).columns[:8]

        # 필터링
        df_trip = df_trip[
            (df_trip['근태항목'] == '관내출장') &
            (df_trip['결재상태'].str.startswith('결재완료'))
        ]

        # 부서 검증
        bepa = ['경영기획실', '청년사업단', '산업인력지원단', '소상공인지원단', '기업지원단', '글로벌사업추진단', '부원장', '기업옴부즈맨실', '임원']
        if not all(dept in bepa for dept in df_trip['부서'].unique()):
            raise ValueError("오류가 발생하였습니다. 개발팀 연락 바랍니다.")

        # 필요한 컬럼 추출
        cols_needed = ['부서', '사원코드', '사원', '직급', '신청일', '시작일', '종료일', '시작시간', '종료시간',
                       '일수', '신청시간', '교통수단', '운전자', '출발지', '도착지', '경유지', '방문처', '목적', '내용']
        df_trip = df_trip[cols_needed].copy()
        
        # 2. 태그 데이터 처리
        df_tag = pd.read_excel(tag_path)
        df_tag = df_tag[['태깅일자', '사원코드', '근태구분', '근무시간']]
        
        tags_out = df_tag[df_tag['근태구분'] == '외출'].sort_values('근무시간').groupby(['태깅일자', '사원코드'])['근무시간'].last().reset_index(name='외출태그')
        tags_in = df_tag[df_tag['근태구분'] == '복귀'].sort_values('근무시간').groupby(['태깅일자', '사원코드'])['근무시간'].first().reset_index(name='복귀태그')

        df_trip = pd.merge(df_trip, tags_out, how='left', left_on=['시작일', '사원코드'], right_on=['태깅일자', '사원코드'])
        df_trip = pd.merge(df_trip, tags_in, how='left', left_on=['시작일', '사원코드'], right_on=['태깅일자', '사원코드'])

        # 3. 로직 적용 함수
        def apply_logic(row):
            str_time = row['시작시간']
            end_time = row['종료시간']
            out_time = row['외출태그']
            in_time = row['복귀태그']
            
            # 기본값: 태그가 있으면 일단 가져옴 (시간 포맷팅)
            if pd.notna(out_time): 
                out_time = str(out_time)[:5]
                out_use = out_time
            else:
                out_use = None
                
            if pd.notna(in_time): 
                in_time = str(in_time)[:5]
                in_use = in_time
            else:
                in_use = None

            # [수정] 로직 1: 신청시간과 태그시간 불일치 -> None (빈칸) 반환
            if pd.notna(out_time) and pd.notna(in_time):
                if (out_time > end_time) or (in_time < str_time):
                    return pd.Series([out_time, in_time, None, None]) 

            # 로직 2: 자동 설정 (없는 경우 인정)
            if (str_time <= '09:00') and pd.isna(out_time):
                out_use = str_time
            elif pd.notna(out_time) and (str_time > out_time):
                out_use = str_time
            
            if (end_time >= '18:00') and pd.isna(in_time):
                in_use = end_time
            elif pd.notna(in_time) and (end_time < in_time):
                in_use = end_time

            return pd.Series([out_time, in_time, out_use, in_use])

        # 로직 적용
        df_trip[['외출태그', '복귀태그', '외출태그(인정)', '복귀태그(인정)']] = df_trip.apply(apply_logic, axis=1)
        
        # 4. 시간 및 여비 계산
        # [수정] calc_out 대신 '외출태그(인정)'을 바로 사용합니다.
        out_dt = pd.to_datetime(df_trip['외출태그(인정)'], format='%H:%M', errors='coerce')
        in_dt = pd.to_datetime(df_trip['복귀태그(인정)'], format='%H:%M', errors='coerce')
        
        diff = in_dt - out_dt
        df_trip['출장시간(산출)/분'] = diff.dt.total_seconds() // 60
        
        def format_duration(x):
            if pd.isna(x): return None
            total_seconds = int(x.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes = remainder // 60
            return f'{hours}:{minutes:02d}'
        
        df_trip['출장시간'] = diff.apply(format_duration)

        # 여비 계산
        conditions = [
            df_trip['출장시간(산출)/분'].isna(), # 인정시간이 None이면 0원
            df_trip['출장시간(산출)/분'] < 240
        ]
        choices = [0, 10000]
        df_trip['여비'] = np.select(conditions, choices, default=20000)
        
        df_trip.loc[df_trip['교통수단'] == '관용차량', '여비'] -= 10000
        df_trip['여비'] = df_trip['여비'].clip(lower=0) 

        # 불필요 컬럼 제거
        df_trip.drop(columns=['태깅일자_x', '사원코드_x', '태깅일자_y', '사원코드_y'], inplace=True, errors='ignore')

    except Exception as e:
        return f"파일 처리 중 오류 발생: {str(e)}"
    
    # 5. 부서별 저장 및 서식 적용
    department_files = []
    final_cols = ['부서', '사원', '직급', '신청일', '시작일', '종료일', '시작시간', 
                  '종료시간', '일수', '신청시간', '외출태그', '복귀태그', '외출태그(인정)', '복귀태그(인정)',
                  '출장시간', '여비', '교통수단', '운전자', '출발지', '도착지', '경유지', '방문처', '목적', '내용']

    for dept, group in df_trip.groupby('부서'):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{dept}_관내여비.xlsx')
        group = group.sort_values(by=['사원', '시작일'] if '시작일' in group.columns else ['사원'])
        group = group[final_cols]
        
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            group.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            max_row = len(group) + 1
            
            # 서식
            gray_bg_format = workbook.add_format({'bg_color': '#D3D3D3'})
            red_text_format = workbook.add_format({'font_color': 'red'})

            worksheet.conditional_format(f'M2:M{max_row}', {'type': 'blanks', 'format': gray_bg_format})
            worksheet.conditional_format(f'N2:N{max_row}', {'type': 'blanks', 'format': gray_bg_format})
            worksheet.conditional_format(f'A2:X{max_row}', {
                'type': 'formula',
                'criteria': '=LEFT($J2, 1)="-"',
                'format': red_text_format
            })
            worksheet.set_column('A:X', 12)

        department_files.append(file_path)

    return render_template('trip_result.html', department_files=department_files)

@app.route('/trip/download/<file_name>')
def download_trip_file(file_name):
    # 업로드 폴더 경로 설정
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return f'파일 {file_name}을 찾을 수 없습니다.'

"""
=============== 교육 담당자 서포터 기능 ===============
"""
@app.route('/edu')
def edu_index():
    return render_template('edu_index.html')

@app.route('/edu/upload', methods=['POST'])
def upload_edu_file():
    if 'file' not in request.files:
        return "파일이 없습니다."

    file = request.files['file']
    if file.filename == '':
        return "파일 이름이 없습니다."

    # CSV 파일 저장
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)

    # CSV 파일 처리
    try:
        df = pd.read_csv(file_path, encoding='CP949', skiprows=1)
    except Exception as e:
        return f"CSV 파일을 처리하는 중 오류가 발생했습니다: {e}"

    df = df.dropna(subset=['연번', '이름']).drop(columns=['비고1', '비고2', 'Unnamed: 14'])
    df = df.rename(columns={'교육\n일시': '교육일시', '교육\n시간': '교육시간', '구분2\n(법정의무/자율)' : '구분2(법정의무/자율)'})
    #df[['연번', '교육시간']] = df[['연번', '교육시간']].astype('int')

    # 체크박스 선택 여부 확인
    include_date = request.form.get('include_date') == 'yes'

    # 중복된 이름이 있는 데이터 추출 (체크박스 선택 시 '교육일시' 포함)
    subset_columns = ['이름', '구분1(외부/내부)', '구분2(법정의무/자율)', '법정 과정', '과정명']
    if include_date:
        subset_columns.append('교육일시')

    duplicated_names = df[df.duplicated(subset=subset_columns, keep=False)]
    duplicated_names = duplicated_names.sort_values(by=['이름', '과정명'])

    # 이름 개수 불일치 확인
    name_counts = df.groupby('과정명')['이름'].agg(고유개수='nunique', 이름개수='count').reset_index()
    name_mismatch = name_counts[name_counts['이름개수'] != name_counts['고유개수']]

    # 공백과 괄호 제거 전 후 비교
    space_yes = df.groupby('과정명').size().reset_index(name='띄어쓰기 제거 전')
    space_yes['과정명'] = space_yes['과정명'].str.replace(' ', '', regex=False)

    bracket_yes = space_yes.groupby('과정명').sum().reset_index()
    bracket_yes.columns = ['과정명', '괄호 제거 전']
    bracket_yes['과정명'] = bracket_yes['과정명'].apply(lambda x:re.sub(r'\(\d+(차시|시간)\)', '', x))

    space_no = df.copy()
    space_no['과정명'] = space_no['과정명'].str.replace(' ', '', regex=False)
    space_no = space_no.groupby('과정명').size().reset_index(name='띄어쓰기 제거 후')

    bracket_no = space_no.copy()
    bracket_no['과정명'] = bracket_no['과정명'].apply(lambda x: re.sub(r'\(\d+(차시|시간)\)', '', x))
    bracket_no = bracket_no.groupby('과정명').sum().reset_index()
    bracket_no.columns = ['과정명', '괄호 제거 후']

    space_df = pd.merge(space_yes, space_no, on='과정명', how='outer')
    space_df['일치 여부'] = space_df['띄어쓰기 제거 전'] == space_df['띄어쓰기 제거 후']
    space_df = space_df[space_df['일치 여부'] == False]
    space_df = space_df.groupby('과정명').agg({
        "띄어쓰기 제거 전" : lambda x: ", ".join(map(str, x)),
        "띄어쓰기 제거 후" : "first"
    }).reset_index()
    space_df.columns = ['과정명', '구분 별 개수', '전체 개수']

    bracket_df = pd.merge(bracket_yes, bracket_no, on='과정명', how='outer')
    bracket_df['일치 여부'] = bracket_df['괄호 제거 전'] == bracket_df['괄호 제거 후']
    bracket_df = bracket_df[bracket_df['일치 여부'] == False]
    bracket_df = bracket_df.groupby('과정명').agg({
        "괄호 제거 전" : lambda x: ", ".join(map(str, x)),
        "괄호 제거 후" : "first"
    }).reset_index()
    bracket_df.columns = ['과정명', '구분 별 개수', '전체 개수']

    # 파일 저장 경로 설정
    duplicated_file = os.path.join(app.config['PROCESSED_FOLDER'], 'duplicated_names.csv')
    mismatch_file = os.path.join(app.config['PROCESSED_FOLDER'], 'name_mismatch.csv')
    space_comparison_file = os.path.join(app.config['PROCESSED_FOLDER'], 'space_comparison.csv')
    bracket_comparison_file = os.path.join(app.config['PROCESSED_FOLDER'], 'bracket_comparison.csv')

    # 처리된 데이터 저장
    duplicated_names.to_csv(duplicated_file, index=False, encoding='CP949')
    name_mismatch[['과정명', '고유개수', '이름개수']].to_csv(mismatch_file, index=False, encoding='CP949')
    space_df.to_csv(space_comparison_file, index=False, encoding='CP949')
    bracket_df.to_csv(bracket_comparison_file, index=False, encoding='CP949')

    # 처리된 데이터프레임을 HTML로 변환하여 보여주기
    return render_template('edu_result.html', 
                           duplicated_names=duplicated_names.to_html(index=False, escape=False),
                           name_mismatch=name_mismatch[['과정명', '고유개수', '이름개수']].to_html(index=False, escape=False),
                           space_comparison=space_df.to_html(index=False, escape=False),
                           bracket_comparison=bracket_df.to_html(index=False, escape=False),
                           duplicated_file='duplicated_names.csv',
                           mismatch_file='name_mismatch.csv',
                           space_comparison_file='space_comparison.csv',
                           bracket_comparison_file='bracket_comparison.csv')

@app.route('/edu/download/<filename>')
def download_edu_file(filename):
    file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
    return send_file(file_path, as_attachment=True)

"""
=============== 신규직원 계정 생성 ===============
"""
def extract_birthdate(jumin):
    if pd.isnull(jumin):
        return None
    jumin = str(jumin).strip().replace('-', '').replace('.', '').split('e')[0]
    if len(jumin) < 7:
        return None
    
    front = jumin[:6]
    gender_code = jumin[6]

    if gender_code in ['1', '2', '5', '6']:
        century = '19'
    elif gender_code in ['3', '4', '7', '8'] :
        century = '20'
    else:
        return None

    return century + front

@app.route('/hr')
def account_index():
    return render_template('hr_index.html')

@app.route('/hr/upload', methods=['POST'])
def upload_and_process_hr_files():
    # 1. 파일 받기
    files = request.files
    f_form = files.get('file_form')    # 1_구글폼 작성 정보.csv
    f_insa = files.get('file_insa')    # 2_기획팀 작성 정보.csv
    f_old = files.get('file_old_form') # 0_사원정보 업데이트 양식.xlsx
    
    if not (f_form and f_insa):
        return "필수 파일이 누락되었습니다."

    # 경로 저장
    form_path = os.path.join(app.config['UPLOAD_FOLDER'], 'hr_form.csv')
    insa_path = os.path.join(app.config['UPLOAD_FOLDER'], 'hr_insa.csv')
    code_path = os.path.join(app.root_path, 'static', 'forms', 'codes.xlsx')
    input_a10_path = os.path.join(app.root_path, 'static', 'forms', 'input_a10.csv')
    input_vpn_path = os.path.join(app.root_path, 'static', 'forms', 'input_vpn.csv')
    old_path = os.path.join(app.config['UPLOAD_FOLDER'], 'hr_old.xlsx')
    
    f_form.save(form_path)
    f_insa.save(insa_path)
    old_path = None
    if f_old and f_old.filename != '':
        old_path = os.path.join(app.config['UPLOAD_FOLDER'], 'hr_old.xlsx')
        f_old.save(old_path)

    try:
        # 2. 데이터 로드 및 전처리
        df_form = pd.read_csv(form_path, dtype='str')
        df_form = df_form.drop(['타임스탬프', '전화번호', 'VPN 계정', '복지카드', '증명사진', '통장사본', '한자 이름'], axis=1, errors='ignore')
        
        df_insa = pd.read_csv(insa_path, dtype='str')

        if os.path.exists(code_path):
            excel_data = pd.read_excel(code_path, sheet_name=None, dtype='string')
            codes = {
                sheet_name: dict(zip(df['항목명'], df['코드']))
                for sheet_name, df in excel_data.items()
            }
        else:
            return "서버에 코드표 파일이 존재하지 않습니다.", 404

        # 데이터 결합
        df = pd.merge(df_form, df_insa, how='outer', on='이름')

        # 데이터 수정 로직
        df['성별'] = df['성별'].apply(lambda x : '여성' if x=='여' else '남성')
        df['주민등록번호'] = df['주민등록번호'].str.replace('-','')
        if '계좌번호' in df.columns:
            df['계좌번호'] = df['계좌번호'].str.replace(' ', '')
            df[['(급여)이체은행', '(급여)계좌번호']] = df['계좌번호'].apply(lambda x : pd.Series(str(x).split('/', 1)) if pd.notna(x) else pd.Series([None, None]))
        
        # 코드 매핑
        mapping_cols = {'팀명': '부서', '(급여)이체은행': '은행', '고용구분': '고용', '직급': '직급', '직책': '직책'}
        for col, code_key in mapping_cols.items():
            if col in df.columns and code_key in codes:
                df[col] = df[col].map(codes[code_key])

        # 3. 파일 생성 1: 상용직 관리 등록
        df_act = pd.read_csv(input_a10_path, dtype='str')
        df_act = pd.DataFrame(index=df.index, columns=df_act.columns)

        df_act['로그인ID'] = df['계정']
        df_act['메일ID'] = df['계정']
        df_act['로그인 비밀번호'] = '111111'
        df_act['프로필명(한국어)'] = df['이름']
        df_act['사원명(한국어)'] = df['이름']
        df_act['성별'] = df['성별']
        df_act['휴대전화'] = df['전화번호']
        df_act['기본주소'] = df['기본주소']
        df_act['사용언어'] = '한국어'
        df_act['기본라이선스'] = 'EBP라이선스'
        df_act['메일'] = 'Y'
        df_act['최초 입사일'] = df['입사일']
        df_act['회사코드'] = '1000'
        df_act['부서코드'] = df['팀명']
        df_act['사번'] = df['사번']
        df_act['직급코드'] = df['직급']
        df_act['직책코드'] = df['직책']
        df_act['재직구분코드'] = 'J01'
        df_act['고용구분코드'] = df['고용구분']
        df_act['직무코드'] = '001'
        df_act['입사일'] = df['입사일']
        df_act['근태사용'] = '사용'
        df_act['조직도'] = '표시'
        df_act['대화/쪽지 조직도'] = '표시'

        template_path = os.path.join(app.root_path, 'static', 'forms', 'template_a10.xlsx')
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], 'A10 상용직관리 업로드.xlsx')

        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active # 첫 번째 시트 선택

            # 기존 샘플 데이터(8행부터) 삭제
            if ws.max_row >= 8:
                ws.delete_rows(8, ws.max_row) 

            # 데이터 입력 (8행부터 시작)
            start_row = 8
            for i, row in df_act.iterrows():
                current_row = start_row + i
                
                # 순번 (A열)
                ws.cell(row=current_row, column=1).value = i + 1
                
                # 데이터 매핑 (엑셀 컬럼 위치에 맞춰 값 입력)
                ws.cell(row=current_row, column=2).value = row['로그인ID']       # B열
                ws.cell(row=current_row, column=3).value = row['메일ID']         # C열
                ws.cell(row=current_row, column=4).value = row['로그인 비밀번호'] # D열
                ws.cell(row=current_row, column=5).value = row['프로필명(한국어)'] # E열
                ws.cell(row=current_row, column=9).value = row['사원명(한국어)']   # I열
                ws.cell(row=current_row, column=13).value = row['성별']           # M열
                ws.cell(row=current_row, column=14).value = row['휴대전화']       # N열
                ws.cell(row=current_row, column=19).value = row['기본주소']       # S열
                ws.cell(row=current_row, column=21).value = row['사용언어']       # U열
                ws.cell(row=current_row, column=26).value = row['기본라이선스']    # W열
                ws.cell(row=current_row, column=27).value = row['메일']           # X열
                ws.cell(row=current_row, column=28).value = row['최초 입사일']     # Z열
                ws.cell(row=current_row, column=30).value = row['회사코드']       # AB열
                ws.cell(row=current_row, column=31).value = row['부서코드']       # AC열
                ws.cell(row=current_row, column=32).value = row['사번']           # AD열
                ws.cell(row=current_row, column=33).value = row['직급코드']       # AE열
                ws.cell(row=current_row, column=34).value = row['직책코드']       # AF열
                ws.cell(row=current_row, column=35).value = row['재직구분코드']    # AG열
                ws.cell(row=current_row, column=36).value = row['고용구분코드']    # AH열
                ws.cell(row=current_row, column=37).value = row['직무코드']       # AI열
                ws.cell(row=current_row, column=42).value = row['입사일']         # AO열
                ws.cell(row=current_row, column=44).value = row['근태사용']       # AQ열
                ws.cell(row=current_row, column=51).value = row['조직도']         # AX열
                ws.cell(row=current_row, column=52).value = row['대화/쪽지 조직도'] # AY열

            wb.save(output_path)
        else:
            df_act.to_excel(output_path)


        # 4. 사원정보 업데이트 파일 생성
        if old_path and os.path.exists(old_path):
            df_old = pd.read_excel(old_path, header=5, dtype=str)
            
            df_update = df.copy()
            df_update['예금주'] = df_update.get('이름') 

            df_new = pd.merge(df_old, df_update, how='left', on='사번', suffixes=('', '_new'))

            for col in df_update.columns:
                if col != '사번' and col in df_old.columns:
                    if f'{col}_new' in df_new.columns:
                        df_new[col] = df_new[col].combine_first(df_new[f'{col}_new'])

            df_new = df_new.drop(columns=[col for col in df_new.columns if col.endswith('_new')])
            df_new = df_new.iloc[:, :22] # 21개 컬럼까지만 사용

            # 추가 가공 로직
            if '주민등록번호' in df_new.columns:
                df_new['생년월일'] = df_new['주민등록번호'].apply(extract_birthdate)
            if '(급여)이체은행' in df_new.columns:
                df_new['(기타)이체은행'] = df_new['(급여)이체은행']
            if '(급여)계좌번호' in df_new.columns:
                df_new['(기타)계좌번호'] = df_new['(급여)계좌번호']
            if '예금주' in df_new.columns:
                df_new['예금주2'] = df_new['예금주']
            if '로그인ID' in df_new.columns:
                df_new['급여이메일'] = df_new['로그인ID'].apply(lambda x: str(x) + '@bepa.kr' if pd.notna(x) else '')
            df_new['직종'] = '001'
            df_new['급여형태'] = '002'

            # 2) [핵심] 엑셀 양식 유지하며 저장하기 (openpyxl 사용)
            update_filename = '사원정보 업데이트 파일.xlsx'
            output_update_path = os.path.join(app.config['PROCESSED_FOLDER'], update_filename)

            # 원본 양식 파일을 복사해서 엽니다
            wb = openpyxl.load_workbook(old_path)
            ws = wb.active

            # 기존 데이터(8행부터)가 있다면 지우고 시작 (헤더인 7행까지는 유지)
            if ws.max_row >= 8:
                ws.delete_rows(8, ws.max_row)

            # 데이터프레임(df_new)의 내용을 8행부터 한 줄씩 입력
            # df_new의 컬럼 순서가 양식의 컬럼 순서와 일치한다고 가정합니다.
            start_row = 8
            # dataframe_to_rows 대신 직접 순회하며 값 입력 (서식 유지에 유리)
            for i, row in df_new.iterrows():
                current_row = start_row + i
                
                ws.cell(row=current_row, column=1).value = i + 1
                ws.cell(row=current_row, column=2).value = row.get('사번')
                ws.cell(row=current_row, column=3).value = row.get('프로필명(한국어)')
                ws.cell(row=current_row, column=4).value = row.get('프로필명(한국어)')
                ws.cell(row=current_row, column=5).value = row.get('로그인ID')
                ws.cell(row=current_row, column=6).value = row.get('회사코드')
                ws.cell(row=current_row, column=7).value = row.get('회사코드')
                ws.cell(row=current_row, column=8).value = row.get('부서코드')
                ws.cell(row=current_row, column=9).value = '000'
                ws.cell(row=current_row, column=10).value = row.get('주민등록번호')
                ws.cell(row=current_row, column=11).value = None
                ws.cell(row=current_row, column=12).value = row.get('생년월일')
                ws.cell(row=current_row, column=13).value = '000'
                ws.cell(row=current_row, column=14).value = row.get('급여이메일')
                ws.cell(row=current_row, column=15).value = row.get('급여형태')
                ws.cell(row=current_row, column=16).value = row.get('직종')
                ws.cell(row=current_row, column=17).value = row.get('(급여)이체은행')
                ws.cell(row=current_row, column=18).value = row.get('(급여)계좌번호')
                ws.cell(row=current_row, column=19).value = row.get('프로필명(한국어)')
                ws.cell(row=current_row, column=20).value = row.get('(기타)이체은행')
                ws.cell(row=current_row, column=21).value = row.get('(기타)계좌번호')
                ws.cell(row=current_row, column=22).value = row.get('프로필명(한국어)')

            ws.delete_rows(8, 1)
            wb.save(output_update_path)
            result_files = [update_filename]
            return render_template('hr_result.html', result_files=result_files)

        # 4. 파일 생성 2: VPN 등록
        df_vpn = pd.read_csv(input_vpn_path)
        df_vpn = pd.DataFrame(index=df.index, columns=df_vpn.columns)

        df_vpn['U_EMAIL'] = df['계정']
        df_vpn['U_NAME'] = df['이름']
        df_vpn['U_JUMINNO'] = 'qwer1234!!'
        df_vpn['U_CN'] = df['팀명']
        df_vpn['U_GROUP'] = 'Default'

        vpn_path = os.path.join(app.config['PROCESSED_FOLDER'], 'VPN 업로드.txt')
        df_vpn.to_csv(vpn_path, index=False, sep=',', encoding='cp949')

        result_files = ['A10 상용직관리 업로드.xlsx', 'VPN 업로드.txt']
        return render_template('hr_result.html', result_files=result_files)

    except Exception as e:
        return f"인사 정보 처리 중 오류 발생: {str(e)}"

@app.route('/hr/download/<file_name>')
def download_hr_file(file_name):
    file_path = os.path.join(app.config['PROCESSED_FOLDER'], file_name)
    return send_file(file_path, as_attachment=True)


"""
=============== 숫자 한글 변환기 ===============
"""
def number_to_korean(num):
    num = int(re.sub(r'[,]', '', num))  # 숫자에서 콤마 제거 후 정수 변환
    
    units = ['', '만', '억', '조', '경']  
    small_units = ['', '십', '백', '천']  
    digits = [''] + list('일이삼사오육칠팔구')  

    if num == 0:
        return '영'
    
    result = []
    unit_index = 0
    
    while num > 0:
        part = num % 10000  
        num //= 10000
        
        if part > 0:
            part_str = ''
            for i in range(4):  
                digit = (part // (10 ** i)) % 10
                if digit != 0:
                    part_str = digits[digit] + small_units[i] + part_str
            
            result.append(part_str + units[unit_index])
        
        unit_index += 1
    
    return ''.join(result[::-1])

@app.route('/money', methods=['GET', 'POST'])
def money_converter():
    input_value = ""
    converted_value = ""
    
    if request.method == 'POST':
        num = request.form.get('number', '')
        try:
            num = re.sub(r'[^0-9]', '', num)  # 숫자만 남기기
            if num:
                input_value = "{:,}".format(int(num))  # 콤마 추가된 입력값
                converted_value = number_to_korean(num)  # 변환값
            else:
                converted_value = "올바른 숫자를 입력하세요."
        except ValueError:
            converted_value = "올바른 숫자를 입력하세요."
    
    return render_template('money.html', input_value=input_value, converted_value=converted_value)

if __name__ == '__main__':
    app.run(debug=True)